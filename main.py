import gradio as gr
import openpyxl
import win32com.client as win32
import pythoncom  # ✅ COM initialization fix
from io import BytesIO

file = "C:\\Users\\t9577ss\\Desktop\\Automation project\\outlook_excel_mailer\\excel_files\\intital dummy file (2).xlsx"

def extract_milestones(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb['Network']
    milestones = sorted(set(cell.value for cell in sheet['A'][1:] if cell.value))
    return milestones

def get_rows(file, milestone):
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb['Network']
    rows = []
    for idx, cell in enumerate(sheet['A'][1:], start=2):
        if cell.value == milestone:
            email = sheet[f'D{idx}'].value
            form_cell = sheet[f'E{idx}']
            display = form_cell.value
            link = form_cell.hyperlink.target if form_cell.hyperlink else "❌ No hyperlink"
            lineitem = sheet[f'B{idx}'].value
            rows.append(f"Row {idx}: {lineitem} → {email} → {display} → {link}")
    return rows

def generate_email(file, milestone, row_text):
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb['Network']
    idx = int(row_text.split()[1].replace(":", ""))

    email = sheet[f'D{idx}'].value
    form_cell = sheet[f'E{idx}']
    form_text = form_cell.value
    form_link = form_cell.hyperlink.target if form_cell.hyperlink else None
    lineitem = sheet[f'B{idx}'].value

    subject = f"Form Reminder - {milestone} | {lineitem}"
    body = f"""<html><body>
        Hi,<br><br>
        Please fill the form for <b>{milestone}</b>, item <b>{lineitem}</b>:<br>
        <a href="{form_link}">{form_text}</a><br><br>
        Thanks,<br>Quality Team
    </body></html>"""

    return email, subject, body, idx

def send_email_outlook(to, subject, html_body):
    try:
        pythoncom.CoInitialize()  # ✅ Ensure COM is initialized
        outlook = win32.Dispatch('Outlook.Application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.Subject = subject
        mail.HTMLBody = html_body
        # Explicitly set the Sent Items folder (optional, for clarity)
        mail.SaveSentMessageFolder = outlook.GetNamespace("MAPI").GetDefaultFolder(5)  # 5 = olFolderSentMail
        mail.Send()
        return f"✅ Email sent to {to}"
    except Exception as e:
        return f"❌ Failed to send email: {e}"

# --- Gradio UI ---
with gr.Blocks() as demo:
    gr.Markdown("## 📧 Outlook Email Sender via Excel (Hyperlink Extractor)")

    with gr.Row():
        excel_file = gr.File(label="📂 Upload Excel (.xlsx)", file_types=[".xlsx"])

    milestone_dropdown = gr.Dropdown(label="🎯 Select Milestone", choices=[], interactive=True)
    row_dropdown = gr.Dropdown(label="📄 Select Row", choices=[], interactive=True)

    email_box = gr.Textbox(label="📬 To (email)", interactive=False)
    subject_box = gr.Textbox(label="📝 Subject", interactive=False)
    body_box = gr.Textbox(label="📨 Email HTML (edit here)", lines=10, interactive=True)
    html_preview = gr.HTML(label="📨 Email Preview (formatted)")

    send_button = gr.Button("🚀 Send Email")
    result_box = gr.Textbox(label="✅ Result", interactive=False)

    def update_milestone_choices(file):
        return gr.update(choices=extract_milestones(file.name), value=None)

    def update_row_choices(file, milestone):
        return gr.update(choices=get_rows(file.name, milestone), value=None)

    def preview_email(file, milestone, row_text):
        email, subject, body, idx = generate_email(file.name, milestone, row_text)
        return email, subject, body, body  # Return HTML for both boxes

    def update_html_preview(body):
        return body  # Just pass through for HTML rendering

    def send_selected_email(file, milestone, row_text, body):
        email, subject, _, idx = generate_email(file.name, milestone, row_text)
        return send_email_outlook(email, subject, body)

    # Bind events
    excel_file.change(update_milestone_choices, inputs=[excel_file], outputs=[milestone_dropdown])
    milestone_dropdown.change(update_row_choices, inputs=[excel_file, milestone_dropdown], outputs=[row_dropdown])
    row_dropdown.change(preview_email, inputs=[excel_file, milestone_dropdown, row_dropdown],
                        outputs=[email_box, subject_box, body_box, html_preview])
    body_box.change(update_html_preview, inputs=[body_box], outputs=[html_preview])
    send_button.click(send_selected_email, inputs=[excel_file, milestone_dropdown, row_dropdown, body_box], outputs=[result_box])

demo.launch()
